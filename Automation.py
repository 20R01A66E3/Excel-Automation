import openpyxl as xl
from openpyxl.chart import BarChart, Reference # type: ignore


def process_sheets(filename, sheet_name, column_to_correct, correction_value, operation_symbol):
    try:
        wb = xl.load_workbook(filename)
        sheet = wb[sheet_name]
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' does not exist in the workbook.")
        return
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
        return

    corrected_col = sheet.max_column + 1

    # Set header for the corrected column
    header_cell = sheet.cell(1, corrected_col)
    header_cell.value = 'Corrected'

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, column_to_correct)
        if cell.value is None:
            print(f"Skipping row {row} as cell is empty.")
            continue
        try:
            value = int(cell.value)
            if operation_symbol == '+':
                corrected_value = value + correction_value
            elif operation_symbol == '-':
                corrected_value = value - correction_value
            elif operation_symbol == '*':
                corrected_value = value * correction_value
            elif operation_symbol == '/':
                if correction_value == 0:
                    print(f"Error: Division by zero at row {row}. Skipping.")
                    continue
                corrected_value = value / correction_value
            else:
                print(f"Unsupported operation: {operation_symbol}")
                return
            correction_cell = sheet.cell(row, corrected_col)
            correction_cell.value = corrected_value
        except ValueError:
            print(f"Non-numeric value in row {row}, column {column_to_correct}. Skipping.")

    wb.save(filename)
    print(f"Processing complete. Changes saved to '{filename}'.")


def create_chart(filename, sheet_name):
    try:
        wb = xl.load_workbook(filename)
        sheet = wb[sheet_name]
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' does not exist in the workbook.")
        return
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
        return
    chart_position = sheet.cell(sheet.max_row + 10, 1)
    try:
        values = Reference(
            sheet, 
            min_row=2, 
            max_row=sheet.max_row, 
            min_col=sheet.max_column, 
            max_col=sheet.max_column
        )
        chart = BarChart()
        chart.add_data(values, titles_from_data=False)
        sheet.add_chart(chart, "A" + str(chart_position.row))
        wb.save(filename)
        print(f"Chart created and saved to '{filename}'.")
    except Exception as e:
        print(f"Error while creating chart: {e}")
